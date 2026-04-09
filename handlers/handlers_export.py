import asyncio
import os
import tempfile
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from typing import List, Tuple
from zoneinfo import ZoneInfo

import openpyxl
from aiogram import Router
from openpyxl.styles import Alignment, Border, Side

from bot import sql, x3
from config import ADMIN_IDS
from logging_config import logger
from aiogram.types import Message, FSInputFile
from aiogram.filters import Command

router = Router()

# Полная видимость длинных полей (payload, transaction id, crypto-идентификаторы и т.п.) в Excel
_EXCEL_COL_WIDTH_MAX = 255


@router.message(Command(commands=['export']))
async def export_database_to_excel(message: Message):
    """Экспорт базы данных в Excel файл"""
    if message.from_user.id not in ADMIN_IDS:
        await message.answer("❌ Эта команда доступна только администраторам.")
        return

    try:
        await message.answer("🔄 Начинаю экспорт базы данных...")

        snapshot = await sql.get_export_snapshot()

        def _sync_build_export() -> str:
            users_list = snapshot["users"]
            payments_list = snapshot["payments"]
            payments_cards_list = snapshot["payments_cards"]
            payments_platega_crypto_list = snapshot["payments_platega_crypto"]
            payments_wata_sbp_list = snapshot["payments_wata_sbp"]
            payments_wata_card_list = snapshot["payments_wata_card"]
            payments_stars_list = snapshot["payments_stars"]
            payments_cryptobot_list = snapshot["payments_cryptobot"]
            gifts_list = snapshot["gifts"]
            online_list = snapshot["online"]
            white_counter_list = snapshot["white_counter"]

            wb = openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # --- Лист USERS ---
            ws_users = wb.create_sheet(title="users")
            users_columns = [
                'id', 'user_id', 'ref', 'is_delete', 'in_panel', 'is_connect',
                'create_user', 'reserve_field', 'subscription_end_date',
                'white_subscription_end_date', 'last_notification_date',
                'last_broadcast_status', 'last_broadcast_date', 'stamp', 'ttclid',
            ]
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))

            # Заголовки
            for col_num, title in enumerate(users_columns, 1):
                cell = ws_users.cell(row=1, column=col_num, value=title)
                cell.alignment = header_alignment
                cell.border = thin_border

            # Данные
            for row_num, user in enumerate(users_list, 2):
                row_data = [
                    user.id, user.user_id, user.ref, user.is_delete,
                    user.in_panel, user.is_connect, user.create_user,
                    user.reserve_field, user.subscription_end_date,
                    user.white_subscription_end_date, user.last_notification_date,
                    user.last_broadcast_status, user.last_broadcast_date,
                    user.stamp, user.ttclid,
                ]
                for col_num, value in enumerate(row_data, 1):
                    # Форматирование дат
                    if col_num in (9, 10, 13) and value:  # subscription_end_date, white_subscription_end_date, last_broadcast_date
                        if isinstance(value, datetime):
                            value = value.strftime('%Y-%m-%d %H:%M:%S')
                    elif col_num == 11 and value:  # last_notification_date
                        if isinstance(value, datetime):
                            value = value.strftime('%Y-%m-%d')
                    cell = ws_users.cell(row=row_num, column=col_num, value=value)
                    cell.border = thin_border

            # Автоширина
            for col in ws_users.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws_users.column_dimensions[col_letter].width = min(max_len + 2, _EXCEL_COL_WIDTH_MAX)

            # --- Лист PAYMENTS (Platega) ---
            ws_payments = wb.create_sheet(title="payments_sbp")
            payments_columns = ['ID', 'User ID', 'Amount', 'Time Created', 'Is Gift', 'Status', 'Transaction_Id', 'payload']
            for col_num, title in enumerate(payments_columns, 1):
                cell = ws_payments.cell(row=1, column=col_num, value=title)
                cell.alignment = header_alignment
                cell.border = thin_border

            for row_num, pay in enumerate(payments_list, 2):
                row_data = [
                    pay.id, pay.user_id, pay.amount, pay.time_created,
                    pay.is_gift, pay.status, pay.transaction_id, pay.payload
                ]
                for col_num, value in enumerate(row_data, 1):
                    if col_num == 4 and value and isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    cell = ws_payments.cell(row=row_num, column=col_num, value=value)
                    cell.border = thin_border

            for col in ws_payments.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws_payments.column_dimensions[col_letter].width = min(max_len + 2, _EXCEL_COL_WIDTH_MAX)



            # --- Лист PAYMENTS_CARDS (платежи по картам) ---
            ws_payments_cards = wb.create_sheet(title="payments_cards")
            cards_columns = ['ID', 'User ID', 'Amount', 'Time Created', 'Is Gift', 'Status', 'Transaction_Id',
                             'Payload']
            for col_num, title in enumerate(cards_columns, 1):
                cell = ws_payments_cards.cell(row=1, column=col_num, value=title)
                cell.alignment = header_alignment
                cell.border = thin_border

            for row_num, pay in enumerate(payments_cards_list, 2):
                row_data = [
                    pay.id, pay.user_id, pay.amount, pay.time_created,
                    pay.is_gift, pay.status, pay.transaction_id, pay.payload
                ]
                for col_num, value in enumerate(row_data, 1):
                    if col_num == 4 and value and isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    cell = ws_payments_cards.cell(row=row_num, column=col_num, value=value)
                    cell.border = thin_border

            for col in ws_payments_cards.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws_payments_cards.column_dimensions[col_letter].width = min(max_len + 2, _EXCEL_COL_WIDTH_MAX)


            # --- Лист PAYMENTS_STARS ---
            ws_payments_stars = wb.create_sheet(title="payments_stars")
            stars_columns = ['ID', 'User ID', 'Amount (Stars)', 'Time Created', 'Is Gift', 'Status', 'payload']
            for col_num, title in enumerate(stars_columns, 1):
                cell = ws_payments_stars.cell(row=1, column=col_num, value=title)
                cell.alignment = header_alignment
                cell.border = thin_border

            for row_num, ps in enumerate(payments_stars_list, 2):
                row_data = [
                    ps.id, ps.user_id, ps.amount, ps.time_created,
                    ps.is_gift, ps.status, ps.payload
                ]
                for col_num, value in enumerate(row_data, 1):
                    if col_num == 4 and value and isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    cell = ws_payments_stars.cell(row=row_num, column=col_num, value=value)
                    cell.border = thin_border

            for col in ws_payments_stars.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws_payments_stars.column_dimensions[col_letter].width = min(max_len + 2, _EXCEL_COL_WIDTH_MAX)

            # --- Лист PAYMENTS_PLATEGA_CRYPTO ---
            ws_platega_crypto = wb.create_sheet(title="payments_platega_crypto")
            platega_crypto_columns = ['ID', 'User ID', 'Amount', 'Time Created', 'Is Gift', 'Status', 'Transaction_Id',
                                      'Payload']
            for col_num, title in enumerate(platega_crypto_columns, 1):
                cell = ws_platega_crypto.cell(row=1, column=col_num, value=title)
                cell.alignment = header_alignment
                cell.border = thin_border

            for row_num, pay in enumerate(payments_platega_crypto_list, 2):
                row_data = [
                    pay.id, pay.user_id, pay.amount, pay.time_created,
                    pay.is_gift, pay.status, pay.transaction_id, pay.payload
                ]
                for col_num, value in enumerate(row_data, 1):
                    if col_num == 4 and value and isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    cell = ws_platega_crypto.cell(row=row_num, column=col_num, value=value)
                    cell.border = thin_border

            for col in ws_platega_crypto.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws_platega_crypto.column_dimensions[col_letter].width = min(max_len + 2, _EXCEL_COL_WIDTH_MAX)

            # --- Лист PAYMENTS_WATA_SBP ---
            ws_wata_sbp = wb.create_sheet(title="payments_wata_sbp")
            wata_sbp_columns = [
                "ID", "User ID", "Amount", "Time Created", "Is Gift", "Status", "Transaction_Id", "Payload"
            ]
            for col_num, title in enumerate(wata_sbp_columns, 1):
                cell = ws_wata_sbp.cell(row=1, column=col_num, value=title)
                cell.alignment = header_alignment
                cell.border = thin_border
            for row_num, pay in enumerate(payments_wata_sbp_list, 2):
                row_data = [
                    pay.id, pay.user_id, pay.amount, pay.time_created,
                    pay.is_gift, pay.status, pay.transaction_id, pay.payload
                ]
                for col_num, value in enumerate(row_data, 1):
                    if col_num == 4 and value and isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    cell = ws_wata_sbp.cell(row=row_num, column=col_num, value=value)
                    cell.border = thin_border
            for col in ws_wata_sbp.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws_wata_sbp.column_dimensions[col_letter].width = min(max_len + 2, _EXCEL_COL_WIDTH_MAX)

            # --- Лист PAYMENTS_WATA_CARD ---
            ws_wata_card = wb.create_sheet(title="payments_wata_card")
            wata_card_columns = [
                "ID", "User ID", "Amount", "Time Created", "Is Gift", "Status", "Transaction_Id", "Payload"
            ]
            for col_num, title in enumerate(wata_card_columns, 1):
                cell = ws_wata_card.cell(row=1, column=col_num, value=title)
                cell.alignment = header_alignment
                cell.border = thin_border
            for row_num, pay in enumerate(payments_wata_card_list, 2):
                row_data = [
                    pay.id, pay.user_id, pay.amount, pay.time_created,
                    pay.is_gift, pay.status, pay.transaction_id, pay.payload
                ]
                for col_num, value in enumerate(row_data, 1):
                    if col_num == 4 and value and isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    cell = ws_wata_card.cell(row=row_num, column=col_num, value=value)
                    cell.border = thin_border
            for col in ws_wata_card.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws_wata_card.column_dimensions[col_letter].width = min(max_len + 2, _EXCEL_COL_WIDTH_MAX)

            # --- Лист PAYMENTS_CRYPTOBOT ---
            ws_payments_cryptobot = wb.create_sheet(title="payments_cryptobot")
            crypto_columns = [
                'ID', 'User ID', 'Amount', 'Currency', 'Time Created',
                'Is Gift', 'Status', 'Invoice ID', 'Payload'
            ]
            for col_num, title in enumerate(crypto_columns, 1):
                cell = ws_payments_cryptobot.cell(row=1, column=col_num, value=title)
                cell.alignment = header_alignment
                cell.border = thin_border

            for row_num, pc in enumerate(payments_cryptobot_list, 2):
                row_data = [
                    pc.id, pc.user_id, pc.amount, pc.currency, pc.time_created,
                    pc.is_gift, pc.status, pc.invoice_id, pc.payload
                ]
                for col_num, value in enumerate(row_data, 1):
                    if col_num == 5 and value and isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    cell = ws_payments_cryptobot.cell(row=row_num, column=col_num, value=value)
                    cell.border = thin_border

            for col in ws_payments_cryptobot.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws_payments_cryptobot.column_dimensions[col_letter].width = min(max_len + 2, _EXCEL_COL_WIDTH_MAX)

            # --- Лист GIFTS ---
            ws_gifts = wb.create_sheet(title="gifts")
            gifts_columns = ['gift_id', 'giver_id', 'duration', 'recepient_id', 'white_flag', 'flag']
            for col_num, title in enumerate(gifts_columns, 1):
                cell = ws_gifts.cell(row=1, column=col_num, value=title)
                cell.alignment = header_alignment
                cell.border = thin_border

            for row_num, gift in enumerate(gifts_list, 2):
                row_data = [
                    gift.gift_id, gift.giver_id, gift.duration,
                    gift.recepient_id, gift.white_flag, gift.flag
                ]
                for col_num, value in enumerate(row_data, 1):
                    cell = ws_gifts.cell(row=row_num, column=col_num, value=value)
                    cell.border = thin_border

            for col in ws_gifts.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws_gifts.column_dimensions[col_letter].width = min(max_len + 2, _EXCEL_COL_WIDTH_MAX)

            # --- Лист ONLINE ---
            ws_online = wb.create_sheet(title="online")
            online_columns = ['ID', 'Дата сбора', 'Всего в панели', 'Активны сегодня', 'Платных', 'Триальных']
            for col_num, title in enumerate(online_columns, 1):
                cell = ws_online.cell(row=1, column=col_num, value=title)
                cell.alignment = header_alignment
                cell.border = thin_border

            for row_num, rec in enumerate(online_list, 2):
                row_data = [
                    rec.online_id, rec.online_date, rec.users_panel,
                    rec.users_active, rec.users_pay, rec.users_trial
                ]
                for col_num, value in enumerate(row_data, 1):
                    if col_num == 2 and value and isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    cell = ws_online.cell(row=row_num, column=col_num, value=value)
                    cell.border = thin_border

            for col in ws_online.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws_online.column_dimensions[col_letter].width = min(max_len + 2, _EXCEL_COL_WIDTH_MAX)

            # --- Лист WHITE_COUNTER ---
            ws_white_counter = wb.create_sheet(title="white_counter")
            wc_columns = ['ID', 'User ID', 'Time Created']
            for col_num, title in enumerate(wc_columns, 1):
                cell = ws_white_counter.cell(row=1, column=col_num, value=title)
                cell.alignment = header_alignment
                cell.border = thin_border

            for row_num, wc in enumerate(white_counter_list, 2):
                row_data = [wc.id, wc.user_id, wc.time_created]
                for col_num, value in enumerate(row_data, 1):
                    if col_num == 3 and value and isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    cell = ws_white_counter.cell(row=row_num, column=col_num, value=value)
                    cell.border = thin_border

            for col in ws_white_counter.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws_white_counter.column_dimensions[col_letter].width = min(max_len + 2, _EXCEL_COL_WIDTH_MAX)

            # Заморозка заголовков
            for ws in [ws_users, ws_payments, ws_payments_cards, ws_payments_stars, ws_platega_crypto,
                       ws_wata_sbp, ws_wata_card, ws_payments_cryptobot, ws_gifts, ws_online, ws_white_counter]:
                ws.freeze_panes = ws['A2']

            fd, path = tempfile.mkstemp(suffix=".xlsx")
            os.close(fd)
            wb.save(path)
            return path

        export_path = await asyncio.to_thread(_sync_build_export)
        users_list = snapshot["users"]
        gifts_list = snapshot["gifts"]
        payments_list = snapshot["payments"]
        payments_cards_list = snapshot["payments_cards"]
        payments_stars_list = snapshot["payments_stars"]
        payments_platega_crypto_list = snapshot["payments_platega_crypto"]
        payments_wata_sbp_list = snapshot["payments_wata_sbp"]
        payments_wata_card_list = snapshot["payments_wata_card"]
        payments_cryptobot_list = snapshot["payments_cryptobot"]

        users_count = len(users_list)
        gifts_count = len(gifts_list)
        payments_count = len(payments_list)
        payments_cards_count = len(payments_cards_list)
        payments_stars_count = len(payments_stars_list)
        payments_cryptobot_count = len(payments_cryptobot_list)
        white_counter_count = len(snapshot["white_counter"])
        payments_platega_crypto_count = len(payments_platega_crypto_list)
        payments_wata_sbp_count = len(payments_wata_sbp_list)
        payments_wata_card_count = len(payments_wata_card_list)
        white_subscription_count = sum(
            1 for u in users_list if u.white_subscription_end_date is not None
        )

        successful_payments_count = sum(1 for p in payments_list if p.status == "confirmed")
        successful_cards_count = sum(1 for p in payments_cards_list if p.status == "confirmed")
        successful_platega_crypto_count = sum(
            1 for p in payments_platega_crypto_list if p.status == "confirmed"
        )
        successful_wata_sbp_count = sum(1 for p in payments_wata_sbp_list if p.status == "confirmed")
        successful_wata_card_count = sum(1 for p in payments_wata_card_list if p.status == "confirmed")
        successful_stars_count = sum(1 for p in payments_stars_list if p.status == "confirmed")
        successful_cryptobot_count = sum(1 for p in payments_cryptobot_list if p.status == "paid")

        try:
            now_s = datetime.now().strftime('%d.%m.%Y %H:%M')
            caption = (
                "📊 Экспорт базы данных\n"
                f"📅 Создано: {now_s}\n\n"
                "📊 Статистика:\n"
                f"├ 👥 Пользователей: {users_count}\n"
                f"├ 🎁 Подарков: {gifts_count}\n"
                f"├ ⚡ Платежей Platega СБП: {successful_payments_count}/{payments_count}\n"
                f"├ 💳 Платежей Platega Карта: {successful_cards_count}/{payments_cards_count}\n"
                f"├ ⭐ Платежей Stars: {successful_stars_count}/{payments_stars_count}\n"
                f"├ 💰 Платежей Platega Крипто: {successful_platega_crypto_count}/{payments_platega_crypto_count}\n"
                f"├ ⚡ Платежей WATA СБП: {successful_wata_sbp_count}/{payments_wata_sbp_count}\n"
                f"├ 💳 Платежей WATA Карта: {successful_wata_card_count}/{payments_wata_card_count}\n"
                f"├ 💎 Платежей Криптоботом: {successful_cryptobot_count}/{payments_cryptobot_count}\n"
                f"├ ⚪ White-подписок: {white_subscription_count}\n"
                f"└ 👁 White-кликов: {white_counter_count}"
            )
            await message.answer_document(
                document=FSInputFile(export_path),
                caption=caption,
            )
        finally:
            try:
                os.remove(export_path)
            except OSError:
                pass

        logger.info(f"Администратор {message.from_user.id} экспортировал базу данных в Excel")

    except Exception as e:
        error_message = f"❌ Ошибка при экспорте базы данных: {str(e)}"
        logger.error(error_message)
        logger.exception("Детали ошибки:")
        await message.answer(error_message)


MSK = ZoneInfo("Europe/Moscow")


def _utc_naive(dt: datetime) -> datetime:
    if dt.tzinfo is not None:
        return dt.astimezone(timezone.utc).replace(tzinfo=None)
    return dt


def _msk_start_as_utc_naive(d: date) -> datetime:
    aware = datetime(d.year, d.month, d.day, tzinfo=MSK)
    return aware.astimezone(timezone.utc).replace(tzinfo=None)


def _payment_msk_date(utc_naive: datetime) -> date:
    return utc_naive.replace(tzinfo=timezone.utc).astimezone(MSK).date()


def _build_billing_xlsx(events: List[Tuple[int, datetime, int]]) -> str:
    """Строит Excel по дням (календарь МСК). events: (user_id, time_created, duration_days)."""
    paid_on_day: dict[date, set[int]] = defaultdict(set)
    first_ts: dict[int, datetime] = {}

    for uid, t_raw, _dur in events:
        t = _utc_naive(t_raw)
        if uid not in first_ts or t < first_ts[uid]:
            first_ts[uid] = t
        paid_on_day[_payment_msk_date(t)].add(uid)

    first_msk_day: dict[int, date] = {
        uid: _payment_msk_date(ts) for uid, ts in first_ts.items()
    }

    events_sorted = sorted(events, key=lambda x: (_utc_naive(x[1]), x[0]))
    today_msk = datetime.now(MSK).date()
    min_day = min(first_msk_day.values()) if first_msk_day else today_msk

    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    headers = [
        "Дата (МСК)",
        "Первобилы сегодня",
        "Рецидивисты сегодня",
        "Всего рецидивистов",
        "Всего рецидивистов, %",
        "Ушедшие",
        "Ушедшие, %",
        "Всего первобилов",
        "Всего первобилов, %",
        "Всего оплативших",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "billing"

    for col_num, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=title)
        cell.alignment = header_alignment
        cell.border = thin_border

    ptr = 0
    user_n: dict[int, int] = {}
    user_end: dict[int, datetime] = {}
    n_rows = 0
    d = min_day
    while d <= today_msk:
        cutoff = _msk_start_as_utc_naive(d + timedelta(days=1))
        while ptr < len(events_sorted):
            uid, t_raw, dur = events_sorted[ptr]
            tn = _utc_naive(t_raw)
            if tn >= cutoff:
                break
            user_n[uid] = user_n.get(uid, 0) + 1
            prev = user_end.get(uid)
            base = max(tn, prev) if prev is not None else tn
            user_end[uid] = base + timedelta(days=dur)
            ptr += 1

        day_users = paid_on_day.get(d, set())
        n_first = sum(1 for u in day_users if first_msk_day.get(u) == d)
        n_repeat = sum(1 for u in day_users if first_msk_day.get(u) is not None and first_msk_day[u] < d)

        total_payers = len(user_n)
        total_recurrent_active = sum(
            1
            for u, c in user_n.items()
            if c >= 2 and user_end.get(u) is not None and user_end[u] >= cutoff
        )
        total_churned = sum(1 for u in user_n if user_end.get(u) is not None and user_end[u] < cutoff)
        total_firstbill_active = sum(
            1
            for u, c in user_n.items()
            if c == 1 and user_end.get(u) is not None and user_end[u] >= cutoff
        )

        pct_recurrent = round(100.0 * total_recurrent_active / total_payers, 2) if total_payers else None
        pct_churned = round(100.0 * total_churned / total_payers, 2) if total_payers else None
        pct_firstbill = round(100.0 * total_firstbill_active / total_payers, 2) if total_payers else None

        row = [
            d.isoformat(),
            n_first,
            n_repeat,
            total_recurrent_active,
            pct_recurrent,
            total_churned,
            pct_churned,
            total_firstbill_active,
            pct_firstbill,
            total_payers,
        ]
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=n_rows + 2, column=col_num, value=value)
            cell.border = thin_border
            if col_num == 1:
                cell.alignment = Alignment(horizontal="center")
        n_rows += 1
        d += timedelta(days=1)

    for idx in range(1, len(headers) + 1):
        letter = openpyxl.utils.get_column_letter(idx)
        max_w = len(str(headers[idx - 1]))
        for row in range(2, n_rows + 2):
            v = ws.cell(row=row, column=idx).value
            if v is not None:
                max_w = max(max_w, len(str(v)))
        ws.column_dimensions[letter].width = min(max_w + 2, _EXCEL_COL_WIDTH_MAX)

    ws.freeze_panes = "A2"
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    return path


@router.message(Command(commands=["billing"]))
async def export_billing_excel(message: Message):
    """Выгрузка метрик по оплатам обычной подписки по дням (Excel)."""
    if message.from_user.id not in ADMIN_IDS:
        return

    await message.answer("🔄 Собираю оплаты обычной подписки и формирую Excel…")
    try:
        events = await sql.get_regular_subscription_payment_events()
        if not events:
            await message.answer(
                "Нет подходящих успешных платежей (confirmed/paid), обычная подписка, не подарок, не mobile, "
                "с известной длительностью (payload или сумма по тарифам)."
            )
            return

        path = await asyncio.to_thread(_build_billing_xlsx, events)
        try:
            fname = f"billing_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            await message.answer_document(
                document=FSInputFile(path, filename=fname),
                caption=(
                    "Оплаты обычной подписки (без «Включи мобильный интернет»), статусы confirmed/paid, "
                    "календарные дни по МСК. Длительность из payload или по сумме (старые платежи без payload)."
                ),
            )
        finally:
            try:
                os.remove(path)
            except OSError:
                pass
        logger.info(f"Администратор {message.from_user.id} выгрузил /billing")
    except Exception as e:
        logger.exception("Ошибка /billing")
        await message.answer(f"❌ Ошибка при выгрузке: {e}")


@router.message(Command("export_panel"))
async def export_panel(message: Message):
    if message.from_user.id not in ADMIN_IDS:
        return

    users_x3 = await x3.get_all_panel()
    total = len(users_x3)
    await message.answer(f"{total} - всего юзеров в панели. Формирую Excel...")

    if not users_x3:
        await message.answer("Нет пользователей для экспорта.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "panel_users"

    # Заголовки
    headers = [
        "username", "telegramId", "expireAt",
        "shortUuid", "vlessUuid", "trojanPassword", "ssPassword",
        "description", "squad_uuid"
    ]
    ws.append(headers)

    # Стили
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Заголовки форматируем
    for col_num, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=title)
        cell.alignment = header_alignment
        cell.border = thin_border

    # Заполнение данными
    for user in users_x3:
        # Извлекаем squad (первый элемент списка activeInternalSquads, если есть)
        squad_name = ""
        squad_uuid = ""
        if user.get('activeInternalSquads') and len(user['activeInternalSquads']) > 0:
            squad = user['activeInternalSquads'][0]
            squad_name = squad.get('name', '')
            squad_uuid = squad.get('uuid', '')

        # Форматируем даты (если есть)
        def format_date(dt_str):
            if dt_str:
                try:
                    dt = datetime.fromisoformat(dt_str.replace('Z', '+00:00'))
                    return dt.strftime('%Y-%m-%d %H:%M:%S')
                except:
                    return dt_str
            return ""

        row_data = [
            user.get('username', ''),
            user.get('telegramId', ''),
            format_date(user.get('expireAt')),
            user.get('shortUuid', ''),
            user.get('vlessUuid', ''),
            user.get('trojanPassword', ''),
            user.get('ssPassword', ''),
            user.get('description', ''),
            squad_uuid
        ]
        ws.append(row_data)

    # Автоширина колонок
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, _EXCEL_COL_WIDTH_MAX)

    # Заморозка заголовка
    ws.freeze_panes = 'A2'

    wb.save('panel.xlsx')

    # Отправляем файл
    from aiogram.types import BufferedInputFile
    await message.answer_document(
        document=FSInputFile('panel.xlsx',
        filename=f"panel_users_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"),
        caption=f"📊 Выгружено пользователей из панели: {total}"
    )

    logger.info(f"Администратор {message.from_user.id} выгрузил список пользователей панели")
